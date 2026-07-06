"""Parse BMsim Google spreadsheet case tabs."""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd

from bmsim.download import download_spreadsheet
from bmsim.submissions import unique_submission_labels
from config import CASE_SHEETS, CACHE_XLSX


@dataclass
class CaseData:
    """Parsed Z-spectra for one benchmark case."""

    case_number: int
    sheet_name: str
    pool_model: str
    prep_description: str
    offset_label: str
    mt_label: str | None
    offsets_ppm: np.ndarray
    submissions: pd.DataFrame  # columns: submission labels, z values per offset
    original_participant_names: list[str]

    @property
    def participant_names(self) -> list[str]:
        return list(self.submissions.columns)

    @property
    def submission_original_names(self) -> dict[str, str]:
        return dict(zip(self.participant_names, self.original_participant_names))


def _parse_offset_value(cell) -> float | None:
    if cell is None:
        return None
    if isinstance(cell, (int, float)):
        return float(cell)
    text = str(cell).strip().replace(" ppm", "")
    try:
        return float(text)
    except ValueError:
        return None


def parse_case_sheet(worksheet) -> CaseData:
    """Parse one openpyxl worksheet (``case N`` tab)."""
    rows = list(worksheet.iter_rows(values_only=True))
    sheet_name = worksheet.title
    case_number = int(sheet_name.replace("case", "").strip())

    pool_model = str(rows[0][1] or "").strip()
    prep_description = str(rows[3][0] or "").strip()
    offset_label = str(rows[4][0] or "").strip()
    mt_label = str(rows[8][1]).strip() if rows[8][1] else None

    header_row = None
    for i, row in enumerate(rows):
        if row and row[0] == "LAB / researcher":
            header_row = i
            break
    if header_row is None:
        raise ValueError(f"No researcher header in sheet {sheet_name}")

    names = []
    col_indices = []
    for j, cell in enumerate(rows[header_row]):
        if j == 0:
            continue
        if cell is None or str(cell).strip() == "":
            continue
        names.append(str(cell).strip())
        col_indices.append(j)

    offsets: list[float] = []
    data_rows: list[list[float]] = []

    for row in rows[header_row + 1 :]:
        offset = _parse_offset_value(row[0])
        if offset is None:
            if offsets:
                break
            continue
        values = []
        has_numeric_value = False
        for j in col_indices:
            val = row[j] if j < len(row) else None
            if val is None or val == "":
                values.append(np.nan)
                continue
            try:
                values.append(float(val))
                has_numeric_value = True
            except (TypeError, ValueError):
                values.append(np.nan)
        if not has_numeric_value:
            if offsets:
                break
            continue
        offsets.append(offset)
        data_rows.append(values)

    if not offsets:
        raise ValueError(f"No spectral data in sheet {sheet_name}")

    z_matrix = np.array(data_rows, dtype=float)
    submission_labels = unique_submission_labels(names)
    submissions = pd.DataFrame(z_matrix, index=offsets, columns=submission_labels)
    submissions.index.name = "offset_ppm"

    return CaseData(
        case_number=case_number,
        sheet_name=sheet_name,
        pool_model=pool_model,
        prep_description=prep_description,
        offset_label=offset_label,
        mt_label=mt_label,
        offsets_ppm=np.array(offsets, dtype=float),
        submissions=submissions,
        original_participant_names=names,
    )


def load_case(
    case_number: int,
    xlsx_path: Path | None = None,
    *,
    download: bool = True,
) -> CaseData:
    path = xlsx_path or CACHE_XLSX
    if download and not path.exists():
        download_spreadsheet(path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet_name = f"case {case_number}"
    if sheet_name not in wb.sheetnames:
        raise KeyError(f"Sheet {sheet_name!r} not in workbook")
    return parse_case_sheet(wb[sheet_name])


def load_all_cases(
    xlsx_path: Path | None = None,
    *,
    download: bool = True,
) -> dict[int, CaseData]:
    path = xlsx_path or CACHE_XLSX
    if download and not path.exists():
        download_spreadsheet(path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    cases = {}
    for sheet_name in CASE_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        case = parse_case_sheet(wb[sheet_name])
        cases[case.case_number] = case
    return cases


def cases_to_long_csv(cases: dict[int, CaseData], path: Path) -> None:
    """Export tidy CSV: case, submission, original participant, offset_ppm, z."""
    frames = []
    for case_num in sorted(cases):
        case = cases[case_num]
        long = case.submissions.stack().reset_index()
        long.columns = ["offset_ppm", "submission", "z"]
        long.insert(0, "case", case_num)
        long.insert(
            2,
            "original_participant",
            long["submission"].map(case.submission_original_names),
        )
        frames.append(long)
    pd.concat(frames, ignore_index=True).to_csv(path, index=False)
